#!/usr/bin/env python3

import argparse
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import string


def json_to_dict(filename):
    """
    Read JSON file and return as dictionary
    """
    try:
        with open(filename) as f:
            data = json.load(f)
            return data
    except Exception as e:
        print(f"Unable to open {filename}: {e}")


def format_sheet(workbook):
    """Iterates over sheets in workbook, formatting cells based on row position."""
    uppercase_alphabet = string.ascii_uppercase
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(name="Arial", size=16, bold=True)
    blue_fill = PatternFill(fill_type="solid", fgColor="0099CCFF")
    grey_fill = PatternFill(fill_type="solid", fgColor="00F1EDED")
    alignment = Alignment(horizontal="left", wrap_text=True)
    regular_font = Font(name="Arial", size=16)
    for sheet in workbook:
        for row in sheet.rows:
            for cell in row:
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = blue_fill
                    cell.border = thin_border
                    cell.alignment = alignment
                else:
                    cell.font = regular_font
                    cell.border = thin_border
                    cell.alignment = alignment
                    # cell.fill = grey_fill

        # Definining dimentions for each column
        if sheet.title == "Progress Report":
            # # column widths for Progress report
            sheet.column_dimensions['A'].width = 60
            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 15
            sheet.column_dimensions['F'].width = 15
            sheet.column_dimensions['G'].width = 15
        else:
            # column widths for each category of ASVS
            sheet.column_dimensions['A'].width = 50
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 100
            sheet.column_dimensions['D'].width = 10
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 50


def create_workbook(json, custom_output_name, custom_columns):
    """Creates workbook based upon ASVS JSON input + (optional) custom naming and column values."""
    wb = Workbook()
    del wb["Sheet"]
    workbook_title = f"{json['ShortName']}-{json['Version']}.xlsx"
    if custom_output_name:
        workbook_title = custom_output_name
    progress = {}
    for requirement in json["Requirements"]:
        sheet_name = f"{requirement['Shortcode']} - {requirement['Name']}"
        progress[sheet_name] = {}
        wb.create_sheet(sheet_name)
        wb.active = wb[sheet_name]
        ws = wb.active

        # Add Data validation for Results
        dv_status = DataValidation(type="list", formula1='"ToDo,Done,NA"', allow_blank=True)
        ws.add_data_validation(dv_status)

        header_list = ["Category", "#", "Description", "Level", "Status", "Comments"] + custom_columns
        ws.append(header_list)
        for category in requirement["Items"]:
            # collect categories, subcategories and no of items in each for preparing the progress report
            progress[sheet_name][f"{category['Shortcode']}: {category['Name']}"] = len(category["Items"])
            for item in category["Items"]:
                ws.append(
                    [
                        category["Name"],
                        item["Shortcode"],
                        item["Description"],
                        int(item["L"]),
                        None,
                        None
                    ]
                )
                # figure out which row we just wrote to
                row = ws.max_row

                # attach the dropdown to just that cell
                dv_status.add(ws[f"E{row}"])

    # create a progress sheet
    sheet_name = "Progress Report"
    wb.create_sheet(sheet_name, index=0)
    wb.active = wb[sheet_name]
    ws = wb.active

    # Adding rows to progress sheet
    header_list = ["Category", "Subcategory", "Total Checks", "ToDo", "Done", "NA", "Progress"]
    ws.append(header_list)
    for category in progress:
        start = 2
        # count the no of items done/todo/na from the individual sheets for each subcategory
        for subcategory, items in progress[category].items():
            countToDo = f"=COUNTIF('{category}'!E${start}:E${start+items-1},$D$1)"
            countDone = f"=COUNTIF('{category}'!E${start}:E${start+items-1},$E$1)"
            countNA = f"=COUNTIF('{category}'!E${start}:E${start+items-1},$F$1)"
            ws.append([category, subcategory, items, countToDo, countDone, countNA, None])
            start += items

    format_sheet(wb)
    wb.save(filename=workbook_title)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-i",
        "--input_file",
        help="Name of ASVS JSON file to parse (e.g. 'OWASP_Application_Security_Verification_Standard_5.0.0_en.json').",
        required="True",
        action="store",
    )
    parser.add_argument(
        "-o",
        "--output_file",
        help="Filename for xlsx output (optional; defaults to ASVS-n.xlsx).",
        default=False,
        action="store",
    )
    parser.add_argument(
        "-c",
        "--columns",
        help="Additional custom column(s) to include in xlsx output, supporting multiple uses (e.g. -c 'Findings').",
        nargs="+",
        action="append",
        default=[],
    )

    args = parser.parse_args()
    args.columns = [val for sublist in args.columns for val in sublist]
    create_workbook(json_to_dict(args.input_file), args.output_file, args.columns)
